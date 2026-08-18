from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Max, Min, Count
from django.utils import timezone
from django.db import OperationalError
from datetime import datetime, timedelta
import json
import re
import jdatetime
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import ChartData, PLCKey, RollPLCData, PLCColumnPreference
from paper.models import Paper
from material.models import Material
from pulp.models import Pulp, pulp_Sampling_Location_names

# Create your views here.

def process_paper_data():
    """
    Process paper data and create chart data points.
    """
    papers = Paper.objects.all().order_by('roll_number')
    created_count = 0
    
    for paper in papers:
        # Process Moisture (humidity)
        if paper.humidity is not None:
            ChartData.objects.get_or_create(
                date=paper.date,
                type='moisture',
                roll_number=paper.roll_number,
                defaults={
                    'value': str(paper.humidity),
                    'start_time': paper.sampling_start_time
                }
            )
            created_count += 1
        
        # Process Burst
        if paper.burst_test:
            # Try to extract numeric value from burst_test field
            burst_match = re.search(r'(\d+\.?\d*)', paper.burst_test)
            if burst_match:
                burst_value = burst_match.group(1)
                ChartData.objects.get_or_create(
                    date=paper.date,
                    type='burst',
                    roll_number=paper.roll_number,
                    defaults={
                        'value': burst_value,
                        'start_time': paper.sampling_start_time
                    }
                )
                created_count += 1
        
        # Process RCT (average of rct1 to rct5)
        rct_values = [paper.rct1, paper.rct2, paper.rct3, paper.rct4, paper.rct5]
        rct_values = [val for val in rct_values if val is not None]
        if rct_values:
            avg_rct = sum(rct_values) / len(rct_values)
            ChartData.objects.get_or_create(
                date=paper.date,
                type='rct',
                roll_number=paper.roll_number,
                defaults={
                    'value': str(round(avg_rct, 2)),
                    'start_time': paper.sampling_start_time
                }
            )
            created_count += 1
        
        # Process MD
        if paper.tensile_strength_md is not None:
            ChartData.objects.get_or_create(
                date=paper.date,
                type='md',
                roll_number=paper.roll_number,
                defaults={
                    'value': str(paper.tensile_strength_md),
                    'start_time': paper.sampling_start_time
                }
            )
            created_count += 1
        
        # Process CD
        if paper.tensile_strength_cd is not None:
            ChartData.objects.get_or_create(
                date=paper.date,
                type='cd',
                roll_number=paper.roll_number,
                defaults={
                    'value': str(paper.tensile_strength_cd),
                    'start_time': paper.sampling_start_time
                }
            )
            created_count += 1
        
        # Process CCT (average of cct1 to cct5)
        cct_values = [paper.cct1, paper.cct2, paper.cct3, paper.cct4, paper.cct5]
        cct_values = [val for val in cct_values if val is not None]
        if cct_values:
            avg_cct = sum(cct_values) / len(cct_values)
            ChartData.objects.get_or_create(
                date=paper.date,
                type='cct',
                roll_number=paper.roll_number,
                defaults={
                    'value': str(round(avg_cct, 2)),
                    'start_time': paper.sampling_start_time
                }
            )
            created_count += 1
        
        # Process GMS (real_grammage)
        if paper.real_grammage is not None:
            ChartData.objects.get_or_create(
                date=paper.date,
                type='gms',
                roll_number=paper.roll_number,
                defaults={
                    'value': str(paper.real_grammage),
                    'start_time': paper.sampling_start_time
                }
            )
            created_count += 1
        
        # Process CUB
        if paper.cub is not None:
            ChartData.objects.get_or_create(
                date=paper.date,
                type='cub',
                roll_number=paper.roll_number,
                defaults={
                    'value': str(paper.cub),
                    'start_time': paper.sampling_start_time
                }
            )
            created_count += 1
    
    return created_count

def process_pulp_data():
    """
    Process pulp data and create chart data points.
    """
    pulps = Pulp.objects.all().order_by('roll_number')
    created_count = 0
    
    for pulp in pulps:
        # Process pH (average of lower and upper pH)
        ph_values = []
        if pulp.lower_ph is not None:
            ph_values.append(pulp.lower_ph)
        if pulp.upper_ph is not None:
            ph_values.append(pulp.upper_ph)
        
        if ph_values:
            avg_ph = sum(ph_values) / len(ph_values)
            # Get date from paper if roll_number exists, otherwise use created_at
            if pulp.roll_number:
                try:
                    paper = Paper.objects.filter(roll_number=str(pulp.roll_number)).first()
                    if paper:
                        date = paper.date
                        start_time = paper.sampling_start_time
                    else:
                        # Convert created_at to jalali date (simplified)
                        date = pulp.created_at.strftime('%Y-%m-%d')
                        start_time = '00:00'
                except:
                    date = pulp.created_at.strftime('%Y-%m-%d')
                    start_time = '00:00'
            else:
                date = pulp.created_at.strftime('%Y-%m-%d')
                start_time = pulp.lower_sampling_time or '00:00'
            
            ChartData.objects.get_or_create(
                date=date,
                type='ph',
                roll_number=str(pulp.roll_number) if pulp.roll_number else 'N/A',
                defaults={
                    'value': str(round(avg_ph, 2)),
                    'start_time': start_time
                }
            )
            created_count += 1
    
    return created_count

@csrf_exempt
@require_http_methods(["GET", "POST"])
def chart_data_api(request):
    """
    API endpoint to get chart data and process new data.
    """
    if request.method == 'POST':
        # Process new data from paper and pulp models
        paper_count = process_paper_data()
        pulp_count = process_pulp_data()
        return JsonResponse({
            'success': True,
            'message': f'Processed {paper_count} paper data points and {pulp_count} pulp data points',
            'paper_count': paper_count,
            'pulp_count': pulp_count
        })
    
    # GET request - return chart data
    # Get all unique roll numbers and sort them numerically
    all_roll_numbers = set()
    
    # Get roll numbers from paper data
    paper_rolls = Paper.objects.values_list('roll_number', flat=True).distinct()
    for roll in paper_rolls:
        all_roll_numbers.add(roll)
    
    # Get roll numbers from pulp data
    pulp_rolls = Pulp.objects.filter(roll_number__isnull=False).values_list('roll_number', flat=True).distinct()
    for roll in pulp_rolls:
        all_roll_numbers.add(str(roll))
    
    # Convert to list and sort numerically
    roll_numbers_list = []
    for roll in all_roll_numbers:
        try:
            # Try to convert to int for proper numeric sorting
            roll_numbers_list.append(int(roll))
        except ValueError:
            # If not numeric, keep as string and add to end
            roll_numbers_list.append(roll)
    
    # Sort numerically first, then alphabetically for non-numeric
    numeric_rolls = [r for r in roll_numbers_list if isinstance(r, int)]
    string_rolls = [r for r in roll_numbers_list if isinstance(r, str)]
    numeric_rolls.sort()
    string_rolls.sort()
    sorted_roll_numbers = [str(r) for r in numeric_rolls] + string_rolls
    
    # Get all chart data
    chart_data = ChartData.objects.all()
    
    # Group data by type and roll number
    series_data = {
        'ph': {},
        'moisture': {},
        'burst': {},
        'rct': {},
        'cct': {},
        'md': {},
        'cd': {},
        'gms': {},
        'cub': {}
    }
    
    # Populate series data with actual values
    for data_point in chart_data:
        if data_point.type in series_data:
            series_data[data_point.type][data_point.roll_number] = {
                'x': data_point.roll_number,
                'y': float(data_point.value) * 25 if data_point.type == 'ph' else float(data_point.value) * 25 if data_point.type == 'moisture' else float(data_point.value) * 5 if data_point.type == 'cd' else float(data_point.value) * 3 if data_point.type == 'md' else float(data_point.value) * 2 if data_point.type == 'gms' else float(data_point.value),
                'rollNumber': data_point.roll_number,
                'samplingTime': data_point.start_time,
                'date': data_point.date,
                'type': 'paper' if data_point.type != 'ph' else 'pulp'
            }
    
    # Create complete series data with null values for missing roll numbers
    complete_series_data = {}
    for type_key, data_dict in series_data.items():
        complete_series_data[type_key] = []
        for roll_number in sorted_roll_numbers:
            if roll_number in data_dict:
                complete_series_data[type_key].append(data_dict[roll_number])
            else:
                # Add null value for missing roll number
                complete_series_data[type_key].append({
                    'x': roll_number,
                    'y': None,
                    'rollNumber': roll_number,
                    'samplingTime': '',
                    'date': '',
                    'type': 'paper' if type_key != 'ph' else 'pulp'
                })
    
    # Convert to chart series format
    series = []
    colors = ['#3B82F6', '#EF4444', '#10B981', '#F59E0B', '#8B5CF6', '#06B6D4', '#F97316', '#84CC16', '#EC4899']
    type_names = {
        'ph': 'pH',
        'moisture': 'Moisture',
        'burst': 'Burst',
        'rct': 'RCT',
        'cct': 'CCT',
        'md': 'MD',
        'cd': 'CD',
        'gms': 'GMS',
        'cub': 'CUB'
    }
    
    for i, (type_key, data) in enumerate(complete_series_data.items()):
        # Always include series, even if all values are null
        series.append({
            'name': type_names[type_key],
            'data': data,
            'color': colors[i]
        })
    
    return JsonResponse({
        'success': True,
        'series': series,
        'roll_numbers': sorted_roll_numbers,
        'total_points': sum(len(data) for data in complete_series_data.values())
    })

@csrf_exempt
@require_http_methods(["GET"])
def clear_chart_data(request):
    """
    Clear all chart data (for testing purposes).
    """
    count = ChartData.objects.count()
    ChartData.objects.all().delete()
    
    return JsonResponse({
        'success': True,
        'message': f'Cleared {count} chart data points'
    })

@csrf_exempt
@require_http_methods(["GET"])
def debug_chart_data(request):
    """
    Debug endpoint to check chart data.
    """
    chart_data = ChartData.objects.all().order_by('roll_number')
    
    debug_info = {
        'total_records': chart_data.count(),
        'sample_records': [],
        'data_types': {},
    }
    
    # Get sample records
    for record in chart_data[:5]:
        debug_info['sample_records'].append({
            'id': record.id,
            'date': record.date,
            'type': record.type,
            'value': record.value,
            'roll_number': record.roll_number,
            'start_time': record.start_time,
        })
    
    # Count by type
    for data_type in ['ph', 'moisture', 'burst', 'rct', 'cct', 'md', 'cd', 'gms', 'cub']:
        debug_info['data_types'][data_type] = chart_data.filter(type=data_type).count()
    
    return JsonResponse(debug_info)

@csrf_exempt
@require_http_methods(["GET"])
def technical_report_data_api(request):
    """
    API endpoint to get technical report data for burst_test, gsm, humidity, and top headbox data.
    """
    # Get time filter parameter
    time_filter = request.GET.get('time_filter', 'daily')
    
    # Calculate date range based on filter
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    now = timezone.now()
    if time_filter == 'daily':
        # Last 7 days
        start_date = now - timedelta(days=7)
    elif time_filter == 'weekly':
        # Last 4 weeks
        start_date = now - timedelta(weeks=4)
    elif time_filter == 'monthly':
        # Last 6 months
        start_date = now - timedelta(days=180)
    else:
        # Default to daily
        start_date = now - timedelta(days=7)
    # Get all unique roll numbers and sort them numerically
    all_roll_numbers = set()
    
    # Get roll numbers from paper data
    paper_rolls = Paper.objects.values_list('roll_number', flat=True).distinct()
    for roll in paper_rolls:
        all_roll_numbers.add(roll)
    
    # Get roll numbers from pulp data
    pulp_rolls = Pulp.objects.filter(roll_number__isnull=False).values_list('roll_number', flat=True).distinct()
    for roll in pulp_rolls:
        all_roll_numbers.add(str(roll))
    
    # Convert to list and sort numerically
    roll_numbers_list = []
    for roll in all_roll_numbers:
        try:
            # Try to convert to int for proper numeric sorting
            roll_numbers_list.append(int(roll))
        except ValueError:
            # If not numeric, keep as string and add to end
            roll_numbers_list.append(roll)
    
    # Sort numerically first, then alphabetically for non-numeric
    numeric_rolls = [r for r in roll_numbers_list if isinstance(r, int)]
    string_rolls = [r for r in roll_numbers_list if isinstance(r, str)]
    numeric_rolls.sort()
    string_rolls.sort()
    sorted_roll_numbers = [str(r) for r in numeric_rolls] + string_rolls
    
    # Get paper data filtered by date range
    papers = Paper.objects.filter(created_at__gte=start_date).order_by('roll_number')
    
    # Get pulp data filtered by date range
    pulps = Pulp.objects.filter(roll_number__isnull=False, created_at__gte=start_date).order_by('roll_number')
    
    # Group data by type and roll number
    series_data = {
        'burst': {},
        'gms': {},
        'moisture': {},
        'upper_headbox_consistency': {},
        'upper_water_filter': {},
        'upper_ph': {},
        'upper_pulp_temperature': {},
        'downpulpcount': {},
        'lower_water_filter': {},
        'lower_ph': {},
        'lower_pulp_temperature': {},
        'tensile_md': {},
        'tensile_cd': {},
        'upper_headbox_consistency_100': {},
        'downpulpcount_100': {},
        'pond8_consistency': {},
        'curtain_consistency': {},
        'thickener_consistency': {}
    }
    
    # Populate series data with paper values
    for paper in papers:
        # Process Burst Test
        if paper.burst_test:
            # Try to extract numeric value from burst_test field
            burst_match = re.search(r'(\d+\.?\d*)', paper.burst_test)
            if burst_match:
                burst_value = float(burst_match.group(1))
                series_data['burst'][paper.roll_number] = {
                    'x': paper.roll_number,
                    'y': burst_value,
                    'rollNumber': paper.roll_number,
                    'samplingStartTime': paper.sampling_start_time,
                    'samplingEndTime': paper.sampling_end_time,
                    'date': paper.date,
                    'type': 'paper'
                }
        
        # Process GSM (real_grammage)
        if paper.real_grammage is not None:
            series_data['gms'][paper.roll_number] = {
                'x': paper.roll_number,
                'y': paper.real_grammage,
                'rollNumber': paper.roll_number,
                'samplingStartTime': paper.sampling_start_time,
                'samplingEndTime': paper.sampling_end_time,
                'date': paper.date,
                'type': 'paper'
            }
        
        # Process Humidity
        if paper.humidity is not None:
            series_data['moisture'][paper.roll_number] = {
                'x': paper.roll_number,
                'y': paper.humidity * 10,
                'rollNumber': paper.roll_number,
                'samplingStartTime': paper.sampling_start_time,
                'samplingEndTime': paper.sampling_end_time,
                'date': paper.date,
                'type': 'paper'
            }
        
        # Process Tensile Strength MD
        if paper.tensile_strength_md is not None:
            series_data['tensile_md'][paper.roll_number] = {
                'x': paper.roll_number,
                'y': paper.tensile_strength_md,
                'rollNumber': paper.roll_number,
                'samplingStartTime': paper.sampling_start_time,
                'date': paper.date,
                'type': 'paper'
            }
        
        # Process Tensile Strength CD
        if paper.tensile_strength_cd is not None:
            series_data['tensile_cd'][paper.roll_number] = {
                'x': paper.roll_number,
                'y': paper.tensile_strength_cd,
                'rollNumber': paper.roll_number,
                'samplingStartTime': paper.sampling_start_time,
                'date': paper.date,
                'type': 'paper'
            }
    
    # Populate series data with pulp values
    for pulp in pulps:
        roll_number = str(pulp.roll_number)
        
        # Get date from paper if roll_number exists, otherwise use created_at
        try:
            paper = Paper.objects.filter(roll_number=roll_number).first()
            if paper:
                date = paper.date
            else:
                # Convert created_at to jalali date (simplified)
                date = pulp.created_at.strftime('%Y-%m-%d')
        except:
            date = pulp.created_at.strftime('%Y-%m-%d')
        
        # Process Upper Headbox Consistency
        if pulp.upper_headbox_consistency is not None:
            series_data['upper_headbox_consistency'][roll_number] = {
                'x': roll_number,
                'y': pulp.upper_headbox_consistency * 100,
                'rollNumber': roll_number,
                'lowerSamplingTime': pulp.lower_sampling_time or '',
                'date': date,
                'type': 'pulp'
            }
        
        # Process Upper Water Filter
        if pulp.upper_water_filter is not None:
            series_data['upper_water_filter'][roll_number] = {
                'x': roll_number,
                'y': pulp.upper_water_filter * 100,
                'rollNumber': roll_number,
                'lowerSamplingTime': pulp.lower_sampling_time or '',
                'date': date,
                'type': 'pulp'
            }
        
        # Process Upper pH
        if pulp.upper_ph is not None:
            series_data['upper_ph'][roll_number] = {
                'x': roll_number,
                'y': pulp.upper_ph * 10,
                'rollNumber': roll_number,
                'lowerSamplingTime': pulp.lower_sampling_time or '',
                'date': date,
                'type': 'pulp'
            }
        
        # Process Upper Pulp Temperature
        if pulp.upper_pulp_temperature is not None:
            series_data['upper_pulp_temperature'][roll_number] = {
                'x': roll_number,
                'y': pulp.upper_pulp_temperature,
                'rollNumber': roll_number,
                'lowerSamplingTime': pulp.lower_sampling_time or '',
                'date': date,
                'type': 'pulp'
            }
        
        # Process Down Pulp Count
        if pulp.downpulpcount is not None:
            series_data['downpulpcount'][roll_number] = {
                'x': roll_number,
                'y': pulp.downpulpcount * 100,
                'rollNumber': roll_number,
                'lowerSamplingTime': pulp.lower_sampling_time or '',
                'date': date,
                'type': 'pulp'
            }
        
        # Process Lower Water Filter
        if pulp.lower_water_filter is not None:
            series_data['lower_water_filter'][roll_number] = {
                'x': roll_number,
                'y': pulp.lower_water_filter * 100,
                'rollNumber': roll_number,
                'lowerSamplingTime': pulp.lower_sampling_time or '',
                'date': date,
                'type': 'pulp'
            }
        
        # Process Lower pH
        if pulp.lower_ph is not None:
            series_data['lower_ph'][roll_number] = {
                'x': roll_number,
                'y': pulp.lower_ph * 10,
                'rollNumber': roll_number,
                'lowerSamplingTime': pulp.lower_sampling_time or '',
                'date': date,
                'type': 'pulp'
            }
        
        # Process Lower Pulp Temperature
        if pulp.lower_pulp_temperature is not None:
            series_data['lower_pulp_temperature'][roll_number] = {
                'x': roll_number,
                'y': pulp.lower_pulp_temperature,
                'rollNumber': roll_number,
                'lowerSamplingTime': pulp.lower_sampling_time or '',
                'date': date,
                'type': 'pulp'
            }
        
        # Process Upper Headbox Consistency * 100
        if pulp.upper_headbox_consistency is not None:
            series_data['upper_headbox_consistency_100'][roll_number] = {
                'x': roll_number,
                'y': pulp.upper_headbox_consistency * 100,
                'rollNumber': roll_number,
                'samplingStartTime': pulp.lower_sampling_time or '',
                'date': date,
                'type': 'pulp'
            }
        
        # Process Down Pulp Count * 100
        if pulp.downpulpcount is not None:
            series_data['downpulpcount_100'][roll_number] = {
                'x': roll_number,
                'y': pulp.downpulpcount * 100,
                'rollNumber': roll_number,
                'samplingStartTime': pulp.lower_sampling_time or '',
                'date': date,
                'type': 'pulp'
            }
        
        # Process Pond 8 Consistency
        if pulp.pond8_consistency is not None:
            series_data['pond8_consistency'][roll_number] = {
                'x': roll_number,
                'y': pulp.pond8_consistency,
                'rollNumber': roll_number,
                'lowerSamplingTime': pulp.lower_sampling_time or '',
                'date': date,
                'type': 'pulp'
            }
        
        # Process Curtain Consistency
        if pulp.curtain_consistency is not None:
            series_data['curtain_consistency'][roll_number] = {
                'x': roll_number,
                'y': pulp.curtain_consistency,
                'rollNumber': roll_number,
                'lowerSamplingTime': pulp.lower_sampling_time or '',
                'date': date,
                'type': 'pulp'
            }
        
        # Process Thickener Consistency
        if pulp.thickener_consistency is not None:
            series_data['thickener_consistency'][roll_number] = {
                'x': roll_number,
                'y': pulp.thickener_consistency,
                'rollNumber': roll_number,
                'lowerSamplingTime': pulp.lower_sampling_time or '',
                'date': date,
                'type': 'pulp'
            }
    
    # Create complete series data with null values for missing roll numbers
    complete_series_data = {}
    for type_key, data_dict in series_data.items():
        complete_series_data[type_key] = []
        for roll_number in sorted_roll_numbers:
            if roll_number in data_dict:
                complete_series_data[type_key].append(data_dict[roll_number])
            else:
                # Add null value for missing roll number
                complete_series_data[type_key].append({
                    'x': roll_number,
                    'y': None,
                    'rollNumber': roll_number,
                    'samplingStartTime': '',
                    'samplingEndTime': '',
                    'lowerSamplingTime': '',
                    'date': '',
                    'type': 'paper' if type_key in ['burst', 'gms', 'moisture'] else 'pulp'
                })
    
    # Convert to chart series format
    series = []
    colors = ['#3B82F6', '#EF4444', '#10B981', '#8B5CF6', '#F59E0B', '#06B6D4', '#F97316', '#84CC16', '#EC4899', '#F59E0B', '#8B5CF6', '#3B82F6', '#FF9800', '#8B5CF6', '#84CC16', '#F97316', '#EC4899', '#06B6D4']  # Different colors for each series
    type_names = {
        'burst': 'تست برست',
        'gms': 'گراماژ',
        'moisture': 'رطوبت',
        'upper_headbox_consistency': 'غلظت هدباکس بالا',
        'upper_water_filter': 'فیلتر آب بالا',
        'upper_ph': 'pH بالا',
        'upper_pulp_temperature': 'دمای خمیر بالا',
        'downpulpcount': 'کانس خمیر پایین',
        'lower_water_filter': 'فیلتر آب پایین',
        'lower_ph': 'pH پایین',
        'lower_pulp_temperature': 'دمای خمیر پایین',
        'tensile_md': 'MD',
        'tensile_cd': 'CD',
        'upper_headbox_consistency_100': 'غلظت هدباکس بالا × 100',
        'downpulpcount_100': 'کانس خمیر پایین × 100',
        'pond8_consistency': 'کانس حوض ۸',
        'curtain_consistency': 'کردان',
        'thickener_consistency': 'تیکنر'
    }
    
    for i, (type_key, data) in enumerate(complete_series_data.items()):
        # Always include series, even if all values are null
        series.append({
            'name': type_names[type_key],
            'data': data,
            'color': colors[i]
        })
    
    return JsonResponse({
        'success': True,
        'series': series,
        'roll_numbers': sorted_roll_numbers,
        'total_points': sum(len(data) for data in complete_series_data.values())
    })


@csrf_exempt
@require_http_methods(["GET"])
def dashboard_stats_api(request):
    """
    API endpoint to get dashboard statistics for daily, weekly, monthly, and overall periods.
    Returns highest/lowest values for grammage, humidity, and headbox fields, plus total production.
    """
    
    def get_shamsi_date_range(period):
        """Get date range for current period in Shamsi calendar"""
        now_jalali = jdatetime.datetime.now()
        
        if period == 'daily':
            # Current day
            start_date = now_jalali.strftime('%Y-%m-%d')
            end_date = start_date
        elif period == 'weekly':
            # Current week (Saturday to Friday in Persian calendar)
            # Get day of week (Saturday = 0, Friday = 6)
            weekday = now_jalali.weekday()  # Monday = 0, Sunday = 6
            # Convert to Persian week (Saturday = 0)
            persian_weekday = (weekday + 2) % 7
            start_of_week = now_jalali - jdatetime.timedelta(days=persian_weekday)
            start_date = start_of_week.strftime('%Y-%m-%d')
            end_date = now_jalali.strftime('%Y-%m-%d')
        elif period == 'monthly':
            # Current month
            start_date = now_jalali.strftime('%Y-%m-01')
            end_date = now_jalali.strftime('%Y-%m-%d')
        else:  # overall
            start_date = None
            end_date = None
        
        return start_date, end_date
    
    def filter_papers_by_period(papers, period):
        """Filter paper records by period"""
        if period == 'overall':
            return papers
        
        start_date, end_date = get_shamsi_date_range(period)
        return papers.filter(date__gte=start_date, date__lte=end_date)
    
    def filter_pulps_by_period(pulps, period):
        """Filter pulp records by period"""
        now = timezone.now()
        
        if period == 'daily':
            start_datetime = now.replace(hour=0, minute=0, second=0, microsecond=0)
            return pulps.filter(created_at__gte=start_datetime)
        elif period == 'weekly':
            # Get start of week (Saturday)
            weekday = now.weekday()  # Monday = 0, Sunday = 6
            persian_weekday = (weekday + 2) % 7
            start_of_week = now - timedelta(days=persian_weekday)
            start_datetime = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
            return pulps.filter(created_at__gte=start_datetime)
        elif period == 'monthly':
            start_datetime = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            return pulps.filter(created_at__gte=start_datetime)
        else:  # overall
            return pulps
    
    def calculate_stats(papers, pulps):
        """Calculate highest/lowest for all metrics"""
        stats = {}
        
        # Paper metrics
        if papers.exists():
            # Grammage
            grammage_values = papers.filter(real_grammage__isnull=False)
            if grammage_values.exists():
                stats['grammage'] = {
                    'highest': grammage_values.aggregate(Max('real_grammage'))['real_grammage__max'],
                    'lowest': grammage_values.aggregate(Min('real_grammage'))['real_grammage__min'],
                }
            else:
                stats['grammage'] = {'highest': None, 'lowest': None}
            
            # Humidity
            humidity_values = papers.filter(humidity__isnull=False)
            if humidity_values.exists():
                stats['humidity'] = {
                    'highest': humidity_values.aggregate(Max('humidity'))['humidity__max'],
                    'lowest': humidity_values.aggregate(Min('humidity'))['humidity__min'],
                }
            else:
                stats['humidity'] = {'highest': None, 'lowest': None}
            
            # BURST (extract numeric value from burst_test field)
            burst_values = []
            for paper in papers.filter(burst_test__isnull=False):
                if paper.burst_test:
                    burst_match = re.search(r'(\d+\.?\d*)', paper.burst_test)
                    if burst_match:
                        try:
                            burst_values.append(float(burst_match.group(1)))
                        except ValueError:
                            pass
            if burst_values:
                stats['burst'] = {
                    'highest': max(burst_values),
                    'lowest': min(burst_values),
                }
            else:
                stats['burst'] = {'highest': None, 'lowest': None}
            
            # COBB
            cobb_values = papers.filter(cub__isnull=False)
            if cobb_values.exists():
                stats['cobb'] = {
                    'highest': cobb_values.aggregate(Max('cub'))['cub__max'],
                    'lowest': cobb_values.aggregate(Min('cub'))['cub__min'],
                }
            else:
                stats['cobb'] = {'highest': None, 'lowest': None}
            
            # MD (Tensile Strength MD)
            md_values = papers.filter(tensile_strength_md__isnull=False)
            if md_values.exists():
                stats['md'] = {
                    'highest': md_values.aggregate(Max('tensile_strength_md'))['tensile_strength_md__max'],
                    'lowest': md_values.aggregate(Min('tensile_strength_md'))['tensile_strength_md__min'],
                }
            else:
                stats['md'] = {'highest': None, 'lowest': None}
            
            # CD (Tensile Strength CD)
            cd_values = papers.filter(tensile_strength_cd__isnull=False)
            if cd_values.exists():
                stats['cd'] = {
                    'highest': cd_values.aggregate(Max('tensile_strength_cd'))['tensile_strength_cd__max'],
                    'lowest': cd_values.aggregate(Min('tensile_strength_cd'))['tensile_strength_cd__min'],
                }
            else:
                stats['cd'] = {'highest': None, 'lowest': None}
            
            # RCT (Average of rct1-5)
            rct_values = []
            for paper in papers:
                paper_rct_values = [paper.rct1, paper.rct2, paper.rct3, paper.rct4, paper.rct5]
                paper_rct_values = [val for val in paper_rct_values if val is not None]
                if paper_rct_values:
                    rct_values.append(sum(paper_rct_values) / len(paper_rct_values))
            if rct_values:
                stats['rct'] = {
                    'highest': max(rct_values),
                    'lowest': min(rct_values),
                }
            else:
                stats['rct'] = {'highest': None, 'lowest': None}
        else:
            stats['grammage'] = {'highest': None, 'lowest': None}
            stats['humidity'] = {'highest': None, 'lowest': None}
            stats['burst'] = {'highest': None, 'lowest': None}
            stats['cobb'] = {'highest': None, 'lowest': None}
            stats['md'] = {'highest': None, 'lowest': None}
            stats['cd'] = {'highest': None, 'lowest': None}
            stats['rct'] = {'highest': None, 'lowest': None}
        
        # Pulp metrics
        if pulps.exists():
            # Lower Wire Stats (آمار توری پایین)
            # Down pulp count (کانس خمیر پایین)
            downpulpcount_values = pulps.filter(downpulpcount__isnull=False)
            if downpulpcount_values.exists():
                stats['downpulpcount'] = {
                    'highest': downpulpcount_values.aggregate(Max('downpulpcount'))['downpulpcount__max'],
                    'lowest': downpulpcount_values.aggregate(Min('downpulpcount'))['downpulpcount__min'],
                }
            else:
                stats['downpulpcount'] = {'highest': None, 'lowest': None}
            
            # Lower water filter (کانس توری پایین)
            lower_water_filter_values = pulps.filter(lower_water_filter__isnull=False)
            if lower_water_filter_values.exists():
                stats['lower_water_filter'] = {
                    'highest': lower_water_filter_values.aggregate(Max('lower_water_filter'))['lower_water_filter__max'],
                    'lowest': lower_water_filter_values.aggregate(Min('lower_water_filter'))['lower_water_filter__min'],
                }
            else:
                stats['lower_water_filter'] = {'highest': None, 'lowest': None}
            
            # Lower headbox freeness (فرینس پایین)
            lower_headbox_freeness_values = pulps.filter(lower_headbox_freeness__isnull=False)
            if lower_headbox_freeness_values.exists():
                stats['lower_headbox_freeness'] = {
                    'highest': lower_headbox_freeness_values.aggregate(Max('lower_headbox_freeness'))['lower_headbox_freeness__max'],
                    'lowest': lower_headbox_freeness_values.aggregate(Min('lower_headbox_freeness'))['lower_headbox_freeness__min'],
                }
            else:
                stats['lower_headbox_freeness'] = {'highest': None, 'lowest': None}
            
            # Lower pH (پایین pH)
            lower_ph_values = pulps.filter(lower_ph__isnull=False)
            if lower_ph_values.exists():
                stats['lower_ph'] = {
                    'highest': lower_ph_values.aggregate(Max('lower_ph'))['lower_ph__max'],
                    'lowest': lower_ph_values.aggregate(Min('lower_ph'))['lower_ph__min'],
                }
            else:
                stats['lower_ph'] = {'highest': None, 'lowest': None}
            
            # Lower pulp temperature (دمای پایین)
            lower_pulp_temperature_values = pulps.filter(lower_pulp_temperature__isnull=False)
            if lower_pulp_temperature_values.exists():
                stats['lower_pulp_temperature'] = {
                    'highest': lower_pulp_temperature_values.aggregate(Max('lower_pulp_temperature'))['lower_pulp_temperature__max'],
                    'lowest': lower_pulp_temperature_values.aggregate(Min('lower_pulp_temperature'))['lower_pulp_temperature__min'],
                }
            else:
                stats['lower_pulp_temperature'] = {'highest': None, 'lowest': None}
            
            # Upper Wire Stats (آمار توری بالا)
            # Upper headbox consistency (کانس توری بالا)
            upper_headbox_consistency_values = pulps.filter(upper_headbox_consistency__isnull=False)
            if upper_headbox_consistency_values.exists():
                stats['upper_headbox_consistency'] = {
                    'highest': upper_headbox_consistency_values.aggregate(Max('upper_headbox_consistency'))['upper_headbox_consistency__max'],
                    'lowest': upper_headbox_consistency_values.aggregate(Min('upper_headbox_consistency'))['upper_headbox_consistency__min'],
                }
            else:
                stats['upper_headbox_consistency'] = {'highest': None, 'lowest': None}
            
            # Upper water filter (کانس توری بالا - filter)
            upper_water_filter_values = pulps.filter(upper_water_filter__isnull=False)
            if upper_water_filter_values.exists():
                stats['upper_water_filter'] = {
                    'highest': upper_water_filter_values.aggregate(Max('upper_water_filter'))['upper_water_filter__max'],
                    'lowest': upper_water_filter_values.aggregate(Min('upper_water_filter'))['upper_water_filter__min'],
                }
            else:
                stats['upper_water_filter'] = {'highest': None, 'lowest': None}
            
            # Upper headbox freeness (فرینس بالا)
            upper_headbox_freeness_values = pulps.filter(upper_headbox_freeness__isnull=False)
            if upper_headbox_freeness_values.exists():
                stats['upper_headbox_freeness'] = {
                    'highest': upper_headbox_freeness_values.aggregate(Max('upper_headbox_freeness'))['upper_headbox_freeness__max'],
                    'lowest': upper_headbox_freeness_values.aggregate(Min('upper_headbox_freeness'))['upper_headbox_freeness__min'],
                }
            else:
                stats['upper_headbox_freeness'] = {'highest': None, 'lowest': None}
            
            # Upper pH (بالا pH)
            upper_ph_values = pulps.filter(upper_ph__isnull=False)
            if upper_ph_values.exists():
                stats['upper_ph'] = {
                    'highest': upper_ph_values.aggregate(Max('upper_ph'))['upper_ph__max'],
                    'lowest': upper_ph_values.aggregate(Min('upper_ph'))['upper_ph__min'],
                }
            else:
                stats['upper_ph'] = {'highest': None, 'lowest': None}
            
            # Upper pulp temperature (دمای بالا)
            upper_pulp_temperature_values = pulps.filter(upper_pulp_temperature__isnull=False)
            if upper_pulp_temperature_values.exists():
                stats['upper_pulp_temperature'] = {
                    'highest': upper_pulp_temperature_values.aggregate(Max('upper_pulp_temperature'))['upper_pulp_temperature__max'],
                    'lowest': upper_pulp_temperature_values.aggregate(Min('upper_pulp_temperature'))['upper_pulp_temperature__min'],
                }
            else:
                stats['upper_pulp_temperature'] = {'highest': None, 'lowest': None}
            
            # Pulp Pool Stats (آمار خمیر حوضها)
            # Pond 8 consistency (حوض 8)
            pond8_consistency_values = pulps.filter(pond8_consistency__isnull=False)
            if pond8_consistency_values.exists():
                stats['pond8_consistency'] = {
                    'highest': pond8_consistency_values.aggregate(Max('pond8_consistency'))['pond8_consistency__max'],
                    'lowest': pond8_consistency_values.aggregate(Min('pond8_consistency'))['pond8_consistency__min'],
                }
            else:
                stats['pond8_consistency'] = {'highest': None, 'lowest': None}
            
            # Curtain consistency (کردان)
            curtain_consistency_values = pulps.filter(curtain_consistency__isnull=False)
            if curtain_consistency_values.exists():
                stats['curtain_consistency'] = {
                    'highest': curtain_consistency_values.aggregate(Max('curtain_consistency'))['curtain_consistency__max'],
                    'lowest': curtain_consistency_values.aggregate(Min('curtain_consistency'))['curtain_consistency__min'],
                }
            else:
                stats['curtain_consistency'] = {'highest': None, 'lowest': None}
            
            # Thickener consistency (تیکنر)
            thickener_consistency_values = pulps.filter(thickener_consistency__isnull=False)
            if thickener_consistency_values.exists():
                stats['thickener_consistency'] = {
                    'highest': thickener_consistency_values.aggregate(Max('thickener_consistency'))['thickener_consistency__max'],
                    'lowest': thickener_consistency_values.aggregate(Min('thickener_consistency'))['thickener_consistency__min'],
                }
            else:
                stats['thickener_consistency'] = {'highest': None, 'lowest': None}
        else:
            # Initialize all stats to None if no pulp data
            stats['downpulpcount'] = {'highest': None, 'lowest': None}
            stats['lower_water_filter'] = {'highest': None, 'lowest': None}
            stats['lower_headbox_freeness'] = {'highest': None, 'lowest': None}
            stats['lower_ph'] = {'highest': None, 'lowest': None}
            stats['lower_pulp_temperature'] = {'highest': None, 'lowest': None}
            stats['upper_headbox_consistency'] = {'highest': None, 'lowest': None}
            stats['upper_water_filter'] = {'highest': None, 'lowest': None}
            stats['upper_headbox_freeness'] = {'highest': None, 'lowest': None}
            stats['upper_ph'] = {'highest': None, 'lowest': None}
            stats['upper_pulp_temperature'] = {'highest': None, 'lowest': None}
            stats['pond8_consistency'] = {'highest': None, 'lowest': None}
            stats['curtain_consistency'] = {'highest': None, 'lowest': None}
            stats['thickener_consistency'] = {'highest': None, 'lowest': None}
        
        # Total production (count of paper records)
        stats['total_production'] = papers.count()
        
        return stats
    
    # Calculate stats for each period
    all_papers = Paper.objects.all()
    all_pulps = Pulp.objects.all()
    
    result = {
        'daily': calculate_stats(
            filter_papers_by_period(all_papers, 'daily'),
            filter_pulps_by_period(all_pulps, 'daily')
        ),
        'weekly': calculate_stats(
            filter_papers_by_period(all_papers, 'weekly'),
            filter_pulps_by_period(all_pulps, 'weekly')
        ),
        'monthly': calculate_stats(
            filter_papers_by_period(all_papers, 'monthly'),
            filter_pulps_by_period(all_pulps, 'monthly')
        ),
        'overall': calculate_stats(
            filter_papers_by_period(all_papers, 'overall'),
            filter_pulps_by_period(all_pulps, 'overall')
        ),
    }
    
    return JsonResponse({
        'success': True,
        'data': result
    })


@csrf_exempt
@require_http_methods(["GET"])
def complete_report_export_xlsx(request):
    """
    Export complete report (paper and pulp data together) to Excel file.
    Matches papers with pulps based on roll_number, date, and time range.
    """
    # Get date range filters (Shamsi dates)
    date_from = request.GET.get('date_from', None)
    date_to = request.GET.get('date_to', None)
    
    # Get other filters
    filter_shift = request.GET.get('shift', None)
    sort_field = request.GET.get('sort_by', '-created_at')
    
    # Get papers with filters
    papers = Paper.objects.all()
    
    if date_from:
        papers = papers.filter(date__gte=date_from)
    if date_to:
        papers = papers.filter(date__lte=date_to)
    if filter_shift:
        papers = papers.filter(shift=filter_shift)
    
    # Apply sorting
    if sort_field.startswith('-'):
        papers = papers.order_by(sort_field[1:]).reverse()
    else:
        papers = papers.order_by(sort_field)
    
    # Get all pulps
    pulps = Pulp.objects.all()
    
    # Get location names for dynamic columns
    location_names = list(pulp_Sampling_Location_names.objects.all().order_by('id'))
    
    # Create material map for material names
    materials = Material.objects.all()
    material_map = {str(material.id): material.material_name for material in materials}
    
    # Helper function to format material usage
    def format_material_usage(material_usage_json):
        if not material_usage_json:
            return ''
        try:
            material_usage = json.loads(material_usage_json)
            formatted_items = []
            for material_id, data in material_usage.items():
                if isinstance(data, dict) and 'val' in data:
                    material_name = material_map.get(material_id, f'Material {material_id}')
                    amount = data.get('val', 0)
                    formatted_items.append(f'{material_name}: {amount}')
            return ', '.join(formatted_items)
        except (json.JSONDecodeError, TypeError):
            return material_usage_json
    
    # Helper function to convert datetime to Jalali
    def to_jalali_datetime(dt):
        if not dt:
            return ''
        try:
            jalali_dt = jdatetime.datetime.fromgregorian(datetime=dt)
            return jalali_dt.strftime('%Y-%m-%d %H:%M:%S')
        except:
            return dt.strftime('%Y-%m-%d %H:%M:%S') if dt else ''
    
    # Helper function to extract date from datetime
    def extract_date(dateTimeStr):
        if not dateTimeStr:
            return ''
        date_part = str(dateTimeStr).split('T')[0].split(' ')[0]
        return date_part
    
    # Helper function to convert time to minutes
    def time_to_minutes(time_str):
        if not time_str:
            return 0
        try:
            parts = str(time_str).split(':')
            hours = int(parts[0]) if len(parts) > 0 else 0
            minutes = int(parts[1]) if len(parts) > 1 else 0
            return hours * 60 + minutes
        except:
            return 0
    
    # Helper function to check if time is between two times
    def is_time_between(time, start_time, end_time):
        time_minutes = time_to_minutes(time)
        start_minutes = time_to_minutes(start_time)
        end_minutes = time_to_minutes(end_time)
        return time_minutes >= start_minutes and time_minutes <= end_minutes
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش کامل"
    
    # Determine PLC keys for export (same visibility logic as frontend)
    pref = PLCColumnPreference.objects.first()
    if pref and pref.visible_keys:
        plc_keys = list(
            PLCKey.objects.filter(id__in=pref.visible_keys).order_by('order_index', 'fa_name')
        )
    else:
        plc_keys = list(PLCKey.objects.all().order_by('order_index', 'fa_name'))

    # Define headers - Paper columns + PLC columns + Pulp columns (18 + dynamic location columns)
    base_paper_headers_before_plc = [
        'نوع رکورد',
        'شماره رول',
        'خط تولید',
        'تاریخ',
        'زمان شروع نمونه‌گیری',
        'زمان پایان نمونه‌گیری',
        'شیفت',
        'نوع کاغذ',
        'عرض کاغذ',
        'گراماژ',
        'رطوبت',
        'خاکستر',
        'کاب',
        'پروفایل',
        'burst',
        'MD',
        'CD',
        'CCT',
        'RCT',
        'پارگی',
    ]
    plc_headers = [pk.fa_name or pk.name or pk.key for pk in plc_keys]
    base_paper_headers_after_plc = [
        'کالندر',
        'سرعت',
        'مواد',
        'دمای سیلندر (قبل/بعد)',
        'دمای کاغذ (نشاسته/پوپ ریل)',
        'غلظت',
        'رقیق‌ساز',
    ]
    paper_headers = base_paper_headers_before_plc + plc_headers + base_paper_headers_after_plc
    
    pulp_headers = [
        'زمان نمونه‌گیری',
        'کانس خمیر پایین',
        'کانس توری پایین',
        'فرینس خمیر پایین',
        'pH پایین',
        'دمای خمیر پایین',
        'کانس خمیر بالا',
        'کانس توری بالا',
        'فرینس خمیر بالا',
        'pH بالا',
        'دمای خمیر بالا',
        'حوض ۸',
        'کردان',
        'تیکنر',
    ]
    
    # Add dynamic location headers
    location_headers = [loc.title for loc in location_names]
    
    # Combine headers
    headers = paper_headers + pulp_headers + location_headers + ['تاریخ ایجاد']
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, size=12, color='FFFFFF')
    
    # Mapping dictionaries
    shift_map = {'day': 'روزانه', 'night': 'شبانه'}
    profile_map = {
        '+1g': '+۱g-',
        '+2g': '+۲g-',
        '+3g': '+۳g-',
        '+4g': '+۴g-',
        '>5g': 'بیشتر از 5 گرم'
    }
    production_line_map = {2: 'PM2', 3: 'PM3', 4: 'PM4', 0: 'مشترک'}
    
    row_num = 2

    # Preload PLC data for rolls
    plc_map = {
        str(obj.roll_number): obj
        for obj in RollPLCData.objects.filter(
            roll_number__in=papers.values_list('roll_number', flat=True)
        )
    }
    
    # Process each paper and its matching pulps
    for paper in papers:
        # Extract paper date from created_at (same as frontend logic)
        paper_date = extract_date(paper.created_at) if paper.created_at else ''
        
        # Find matching pulps
        matching_pulps = []
        for pulp in pulps:
            # Check roll_number match
            if str(pulp.roll_number) != str(paper.roll_number):
                continue
            
            # Check date match (same created_at date)
            pulp_date = extract_date(pulp.created_at)
            if pulp_date != paper_date:
                continue
            
            # Check time range
            if (pulp.lower_sampling_time and 
                paper.sampling_start_time and 
                paper.sampling_end_time):
                if is_time_between(pulp.lower_sampling_time, 
                                   paper.sampling_start_time, 
                                   paper.sampling_end_time):
                    matching_pulps.append(pulp)
        
        # Write paper row
        col_num = 1
        ws.cell(row=row_num, column=col_num, value='کاغذ'); col_num += 1
        ws.cell(row=row_num, column=col_num, value=paper.roll_number or ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=production_line_map.get(paper.ProductionLine, '') if paper.ProductionLine else ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=paper.date or ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=paper.sampling_start_time or ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=paper.sampling_end_time or ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=shift_map.get(paper.shift, '') if paper.shift else ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=paper.PaperType.name if paper.PaperType else ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=paper.paper_size or ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=paper.real_grammage or ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=paper.humidity or ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=paper.ash_percentage or ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=paper.cub or ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=profile_map.get(paper.profile, paper.profile or '') if paper.profile else ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=paper.burst_test or ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=paper.tensile_strength_md or ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=paper.tensile_strength_cd or ''); col_num += 1
        
        # CCT (combine all values)
        cct_values = [paper.cct1, paper.cct2, paper.cct3, paper.cct4, paper.cct5]
        cct_values = [str(v) for v in cct_values if v is not None]
        ws.cell(row=row_num, column=col_num, value='\n'.join(cct_values) if cct_values else ''); col_num += 1
        
        # RCT (combine all values)
        rct_values = [paper.rct1, paper.rct2, paper.rct3, paper.rct4, paper.rct5]
        rct_values = [str(v) for v in rct_values if v is not None]
        ws.cell(row=row_num, column=col_num, value='\n'.join(rct_values) if rct_values else ''); col_num += 1
        
        ws.cell(row=row_num, column=col_num, value=paper.NumberOfTears or ''); col_num += 1

        # PLC-derived fields (from RollPLCData), one column per PLC key
        roll_plc = plc_map.get(str(paper.roll_number))
        for pk in plc_keys:
            val = None
            if roll_plc:
                if pk.key == 'b':
                    val = roll_plc.paper_breaks
                elif pk.key == 'me1':
                    val = roll_plc.printed_length
                else:
                    setting = roll_plc.plc_setting or {}
                    val = setting.get(pk.key)
            ws.cell(row=row_num, column=col_num, value=val if (val is not None and val != '') else '')
            col_num += 1

        ws.cell(row=row_num, column=col_num, value='بله' if paper.calender_applied else 'خیر'); col_num += 1
        ws.cell(row=row_num, column=col_num, value=paper.machine_speed or ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=format_material_usage(paper.material_usage)); col_num += 1
        # Get temperature values from PM_Setting
        pm_settings = paper.pm_settings.all()
        temp_before_values = [str(s.cylinder_temperature_before_press) for s in pm_settings if s.cylinder_temperature_before_press is not None]
        temp_after_values = [str(s.cylinder_temperature_after_press) for s in pm_settings if s.cylinder_temperature_after_press is not None]
        temp_display = f"{', '.join(temp_before_values) if temp_before_values else '-'} / {', '.join(temp_after_values) if temp_after_values else '-'}"
        ws.cell(row=row_num, column=col_num, value=temp_display); col_num += 1
        # Paper temperature values
        paper_temp_starch_values = [str(s.paper_temperature_before_starch) for s in pm_settings if s.paper_temperature_before_starch is not None]
        paper_temp_pop_reel_values = [str(s.paper_temperature_before_pop_reel) for s in pm_settings if s.paper_temperature_before_pop_reel is not None]
        paper_temp_display = f"{', '.join(paper_temp_starch_values) if paper_temp_starch_values else '-'} / {', '.join(paper_temp_pop_reel_values) if paper_temp_pop_reel_values else '-'}"
        ws.cell(row=row_num, column=col_num, value=paper_temp_display); col_num += 1
        ws.cell(row=row_num, column=col_num, value=paper.density_valve or ''); col_num += 1
        ws.cell(row=row_num, column=col_num, value=paper.diluting_valve or ''); col_num += 1
        
        # Pulp columns (empty for paper row)
        for _ in range(len(pulp_headers) + len(location_headers) + 1):
            col_num += 1
        
        row_num += 1
        
        # Write matching pulp rows
        for pulp in matching_pulps:
            col_num = 1
            ws.cell(row=row_num, column=col_num, value='خمیر'); col_num += 1
            ws.cell(row=row_num, column=col_num, value=pulp.roll_number or ''); col_num += 1
            ws.cell(row=row_num, column=col_num, value=production_line_map.get(pulp.ProductionLine, '') if pulp.ProductionLine else ''); col_num += 1
            
            # Paper columns (empty for pulp row)
            for _ in range(len(paper_headers) - 3):  # Skip paper-specific columns except type/roll/line
                col_num += 1
            
            # Pulp columns
            ws.cell(row=row_num, column=col_num, value=pulp.lower_sampling_time or ''); col_num += 1
            ws.cell(row=row_num, column=col_num, value=pulp.downpulpcount or ''); col_num += 1
            ws.cell(row=row_num, column=col_num, value=pulp.lower_water_filter or ''); col_num += 1
            ws.cell(row=row_num, column=col_num, value=pulp.lower_headbox_freeness or ''); col_num += 1
            ws.cell(row=row_num, column=col_num, value=pulp.lower_ph or ''); col_num += 1
            ws.cell(row=row_num, column=col_num, value=pulp.lower_pulp_temperature or ''); col_num += 1
            ws.cell(row=row_num, column=col_num, value=pulp.upper_headbox_consistency or ''); col_num += 1
            ws.cell(row=row_num, column=col_num, value=pulp.upper_water_filter or ''); col_num += 1
            ws.cell(row=row_num, column=col_num, value=pulp.upper_headbox_freeness or ''); col_num += 1
            ws.cell(row=row_num, column=col_num, value=pulp.upper_ph or ''); col_num += 1
            ws.cell(row=row_num, column=col_num, value=pulp.upper_pulp_temperature or ''); col_num += 1
            ws.cell(row=row_num, column=col_num, value=pulp.pond8_consistency or ''); col_num += 1
            ws.cell(row=row_num, column=col_num, value=pulp.curtain_consistency or ''); col_num += 1
            ws.cell(row=row_num, column=col_num, value=pulp.thickener_consistency or ''); col_num += 1
            
            # Dynamic location columns
            for loc in location_names:
                location_value = pulp.sampling_locations.filter(title=loc.title).first()
                ws.cell(row=row_num, column=col_num, value=location_value.value if location_value else ''); col_num += 1
            
            ws.cell(row=row_num, column=col_num, value=to_jalali_datetime(pulp.created_at)); col_num += 1
            
            row_num += 1
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if row.value:
                    max_length = max(max_length, len(str(row.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Build filename with date range
    if date_from and date_to:
        filename = f'complete-report-{date_from}-{date_to}.xlsx'
    elif date_from:
        filename = f'complete-report-{date_from}-.xlsx'
    elif date_to:
        filename = f'complete-report--{date_to}.xlsx'
    else:
        filename = f'complete-report-{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # Set Content-Disposition header with proper encoding
    response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}'
    
    wb.save(response)
    return response


PLC_ROLLS_API_URL = "http://192.168.2.46:6010/api/rolls/?time_range=7"


def _fetch_plc_api(time_range=None):
    """
    Fetch data from external PLC rolls API.
    """
    params = {}
    # if time_range is not None:
    #     params['time_range'] = time_range

    try:
        response = requests.post(PLC_ROLLS_API_URL, params=params, timeout=10)
    except requests.RequestException as exc:
        return None, {'success': False, 'error': str(exc)}

    if response.status_code != 200:
        return None, {
            'success': False,
            'status_code': response.status_code,
            'error': response.text,
        }

    try:
        payload = response.json()
    except json.JSONDecodeError:
        return None, {'success': False, 'error': 'Invalid JSON response from PLC API'}

    return payload, None


def _sync_plc_data_from_payload(payload):
    """
    Upsert PLCKey and RollPLCData instances from API payload.
    """
    plc_keys_data = payload.get('plc_keys', []) or []
    rolls_data = payload.get('data', []) or []

    created_keys = 0
    updated_keys = 0

    for item in plc_keys_data:
        external_id = item.get('id')
        if external_id is None:
            continue

        defaults = {
            'name': item.get('name') or '',
            'fa_name': item.get('fa_name') or '',
            'key': item.get('key') or '',
            'value_type': item.get('value') or '',
            'order_index': item.get('order_index') or 0,
            'description': item.get('description') or '',
            'creation_datetime': item.get('CreationDateTime'),
            'last_update': item.get('LastUpdate'),
        }

        obj, created = PLCKey.objects.update_or_create(
            external_id=external_id,
            defaults=defaults,
        )
        if created:
            created_keys += 1
        else:
            updated_keys += 1

    created_rolls = 0
    updated_rolls = 0

    for item in rolls_data:
        roll_number = item.get('roll_number')
        if roll_number is None:
            continue

        plc_setting = item.get('plc_setting') or {}

        defaults = {
            'plc_setting': plc_setting,
            'creation_datetime': item.get('CreationDateTime'),
            'paper_breaks': item.get('Paper_breaks'),
            'printed_length': item.get('Printed_length'),
        }

        obj, created = RollPLCData.objects.update_or_create(
            roll_number=str(roll_number),
            defaults=defaults,
        )
        if created:
            created_rolls += 1
        else:
            updated_rolls += 1

    return {
        'plc_keys': {
            'created': created_keys,
            'updated': updated_keys,
            'total': PLCKey.objects.count(),
        },
        'roll_plc_data': {
            'created': created_rolls,
            'updated': updated_rolls,
            'total': RollPLCData.objects.count(),
        },
    }


@csrf_exempt
@require_http_methods(["POST", "GET"])
def sync_plc_data_api(request):
    """
    API endpoint to sync PLC data from external system.
    """
    time_range = request.GET.get('time_range') or request.POST.get('time_range')

    if RollPLCData.objects.exists() and not time_range:
        time_range = 1

    payload, error = _fetch_plc_api(time_range=time_range)
    if error is not None:
        return JsonResponse(error, status=502)
    try:
        result = _sync_plc_data_from_payload(payload)
    except OperationalError as exc:
        return JsonResponse(
            {
                'success': False,
                'error': 'database_locked',
                'message': str(exc),
            },
            status=503,
        )

    return JsonResponse({
        'success': True,
        'time_range': time_range,
        'sync_result': result,
    })


@csrf_exempt
@require_http_methods(["GET"])
def plc_keys_api(request):
    """
    API endpoint to get PLC keys metadata.
    """
    keys = list(
        PLCKey.objects.all()
        .order_by('order_index', 'fa_name')
        .values('id', 'name', 'fa_name', 'key', 'value_type', 'order_index', 'description')
    )

    return JsonResponse({
        'success': True,
        'plc_keys': keys,
    })


@csrf_exempt
@require_http_methods(["GET"])
def roll_plc_data_api(request):
    """
    API endpoint to get roll PLC data from local database.
    """
    qs = RollPLCData.objects.all()

    roll_numbers_param = request.GET.get('roll_numbers')
    if roll_numbers_param:
        roll_numbers = [
            rn.strip() for rn in str(roll_numbers_param).split(',') if rn.strip()
        ]
        if roll_numbers:
            qs = qs.filter(roll_number__in=roll_numbers)

    data = list(
        qs.values('roll_number', 'plc_setting', 'creation_datetime', 'paper_breaks', 'printed_length')
    )

    return JsonResponse({
        'success': True,
        'data': data,
    })


@csrf_exempt
@require_http_methods(["GET", "POST"])
def plc_column_preference_api(request):
    """
    API endpoint to get/set global PLC column visibility preferences.
    """
    pref, _ = PLCColumnPreference.objects.get_or_create(id=1)

    if request.method == 'GET':
        return JsonResponse({
            'success': True,
            'visible_keys': pref.visible_keys or [],
        })

    try:
        body = request.body.decode('utf-8') if request.body else '{}'
        data = json.loads(body or '{}')
    except json.JSONDecodeError:
        data = {}

    visible_keys = data.get('visible_keys') or []
    normalized_ids = []
    for value in visible_keys:
        try:
            normalized_ids.append(int(value))
        except (TypeError, ValueError):
            continue

    pref.visible_keys = normalized_ids
    pref.save()

    return JsonResponse({
        'success': True,
        'visible_keys': pref.visible_keys,
    })

