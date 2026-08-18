"""
Views for paper app.
"""
import json
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from django.db.models import Q
from django.http import HttpResponse
from .models import Paper, ProductionMachine, PM_Setting
from .serializers import PaperSerializer, PaperListSerializer, ProductionMachineSerializer, PM_SettingSerializer
from logs.utils import log_action
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import jdatetime
from material.models import Material


class PaperViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Paper model with CRUD operations.
    """
    queryset = Paper.objects.prefetch_related('pm_settings', 'pm_settings__production_machine').all()
    serializer_class = PaperSerializer
    permission_classes = [AllowAny]
    
    def get_serializer_class(self):
        """
        Return appropriate serializer based on action.
        """
        if self.action == 'list':
            return PaperListSerializer
        return PaperSerializer
    
    def get_queryset(self):
        """
        Filter queryset based on query parameters.
        """
        queryset = Paper.objects.prefetch_related('pm_settings', 'pm_settings__production_machine').all()
        
        # Search functionality
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(roll_number__icontains=search) |
                Q(responsible_person_name__icontains=search) |
                Q(date__icontains=search)
            )
        
        # Filter by shift
        shift = self.request.query_params.get('shift', None)
        if shift:
            queryset = queryset.filter(shift=shift)
        
        # Filter by production line
        production_line = self.request.query_params.get('ProductionLine', None)
        if production_line:
            queryset = queryset.filter(ProductionLine=production_line)
        
        # Filter by paper type (ForeignKey)
        paper_type = self.request.query_params.get('paper_type', None)
        if paper_type:
            queryset = queryset.filter(PaperType_id=paper_type)
        
        # Sorting
        sort_by = self.request.query_params.get('sort_by', '-created_at')
        if sort_by:
            queryset = queryset.order_by(sort_by)
        
        return queryset
    
    def perform_create(self, serializer):
        """
        Create paper record and log the action.
        """
        paper = serializer.save()
        
        # Handle PM_Setting data if provided
        pm_settings_data = self.request.data.get('pm_settings', [])
        if pm_settings_data:
            for setting_data in pm_settings_data:
                production_machine_id = setting_data.get('production_machine')
                if production_machine_id:
                    PM_Setting.objects.update_or_create(
                        paper=paper,
                        production_machine_id=production_machine_id,
                        defaults={
                            'bottom': setting_data.get('bottom', ''),
                            'top': setting_data.get('top', ''),
                            'cylinder_temperature_before_press': setting_data.get('cylinder_temperature_before_press'),
                            'cylinder_temperature_after_press': setting_data.get('cylinder_temperature_after_press')
                        }
                    )
        
        # Log action if user is authenticated
        if self.request.user.is_authenticated:
            try:
                log_action(self.request.user.username, 'Paper', 'create')
            except:
                pass
    
    def perform_update(self, serializer):
        """
        Update paper record and log the action.
        """
        # Save the updated instance
        updated_paper = serializer.save()
        
        # Handle PM_Setting data if provided
        pm_settings_data = self.request.data.get('pm_settings', [])
        if pm_settings_data:
            for setting_data in pm_settings_data:
                production_machine_id = setting_data.get('production_machine')
                if production_machine_id:
                    PM_Setting.objects.update_or_create(
                        paper=updated_paper,
                        production_machine_id=production_machine_id,
                        defaults={
                            'bottom': setting_data.get('bottom', ''),
                            'top': setting_data.get('top', ''),
                            'cylinder_temperature_before_press': setting_data.get('cylinder_temperature_before_press'),
                            'cylinder_temperature_after_press': setting_data.get('cylinder_temperature_after_press')
                        }
                    )
        
        # Log action if user is authenticated
        if self.request.user.is_authenticated:
            try:
                log_action(self.request.user.username, 'Paper', 'edit')
            except:
                pass
    
    def perform_destroy(self, instance):
        """
        Delete paper record and log the action.
        """
        instance.delete()
        
        # Log action if user is authenticated
        if self.request.user.is_authenticated:
            try:
                log_action(self.request.user.username, 'Paper', 'delete')
            except:
                pass
    
    @action(detail=False, methods=['get'])
    def suggestions(self, request):
        """
        Get suggestions for autocomplete fields.
        """
        # Get unique values for suggestions with case-insensitive deduplication
        responsible_persons_raw = Paper.objects.values_list('responsible_person_name', flat=True).distinct()
        shifts_raw = Paper.objects.values_list('shift', flat=True).distinct()
        material_usage_raw = Paper.objects.values_list('material_usage', flat=True).distinct()
        machine_speed_raw = Paper.objects.values_list('machine_speed', flat=True).distinct()
        paper_size_raw = Paper.objects.values_list('paper_size', flat=True).distinct()
        
        # Get temperature suggestions from PM_Setting model
        temp_before_press_raw = PM_Setting.objects.values_list('cylinder_temperature_before_press', flat=True).distinct()
        temp_after_press_raw = PM_Setting.objects.values_list('cylinder_temperature_after_press', flat=True).distinct()
        
        # Filter out empty values and normalize for deduplication
        responsible_persons_clean = []
        shifts_clean = []
        material_usage_suggestions = {}
        temp_before_press_clean = []
        temp_after_press_clean = []
        machine_speed_clean = []
        paper_size_clean = []
        
        # Deduplicate responsible person names
        seen_names = set()
        for name in responsible_persons_raw:
            if name and name.strip():  # Check if name exists and is not just whitespace
                normalized_name = name.strip()  # Remove leading/trailing whitespace
                if normalized_name.lower() not in seen_names:  # Case-insensitive check
                    seen_names.add(normalized_name.lower())
                    responsible_persons_clean.append(normalized_name)
        
        # Deduplicate shifts
        seen_shifts = set()
        for shift in shifts_raw:
            if shift and shift.strip():
                normalized_shift = shift.strip()
                if normalized_shift.lower() not in seen_shifts:
                    seen_shifts.add(normalized_shift.lower())
                    shifts_clean.append(normalized_shift)
        
        # Deduplicate temperature before press values
        seen_temp_before = set()
        for temp in temp_before_press_raw:
            if temp is not None:
                if temp not in seen_temp_before:
                    seen_temp_before.add(temp)
                    temp_before_press_clean.append(temp)
        
        # Deduplicate temperature after press values
        seen_temp_after = set()
        for temp in temp_after_press_raw:
            if temp is not None:
                if temp not in seen_temp_after:
                    seen_temp_after.add(temp)
                    temp_after_press_clean.append(temp)
        
        # Deduplicate machine speed values
        seen_machine_speed = set()
        for speed in machine_speed_raw:
            if speed is not None:
                if speed not in seen_machine_speed:
                    seen_machine_speed.add(speed)
                    machine_speed_clean.append(speed)
        
        # Deduplicate paper size values
        seen_paper_size = set()
        for size in paper_size_raw:
            if size is not None:
                if size not in seen_paper_size:
                    seen_paper_size.add(size)
                    paper_size_clean.append(size)
        
        # Process material usage suggestions
        for material_usage_str in material_usage_raw:
            if material_usage_str and material_usage_str.strip():
                try:
                    # Parse JSON material usage data
                    material_data = json.loads(material_usage_str)
                    for material_id, data in material_data.items():
                        if isinstance(data, dict) and 'val' in data and 'brand' in data:
                            if material_id not in material_usage_suggestions:
                                material_usage_suggestions[material_id] = {
                                    'amounts': set(),
                                    'brands': set()
                                }
                            
                            # Add amount suggestion
                            if data['val'] is not None:
                                material_usage_suggestions[material_id]['amounts'].add(data['val'])
                            
                            # Add brand suggestion
                            if data['brand'] and data['brand'].strip():
                                material_usage_suggestions[material_id]['brands'].add(data['brand'].strip())
                                
                except (json.JSONDecodeError, TypeError):
                    # Handle old format or invalid JSON
                    continue
        
        # Convert sets to sorted lists for JSON serialization
        for material_id in material_usage_suggestions:
            material_usage_suggestions[material_id]['amounts'] = sorted(list(material_usage_suggestions[material_id]['amounts']))
            material_usage_suggestions[material_id]['brands'] = sorted(list(material_usage_suggestions[material_id]['brands']))
        
        # Sort alphabetically and numerically
        responsible_persons_clean.sort()
        shifts_clean.sort()
        temp_before_press_clean.sort()
        temp_after_press_clean.sort()
        machine_speed_clean.sort()
        paper_size_clean.sort()
        
        return Response({
            'responsible_person_names': responsible_persons_clean,
            'shifts': shifts_clean,
            'temp_before_press_suggestions': temp_before_press_clean,
            'temp_after_press_suggestions': temp_after_press_clean,
            'machine_speed_suggestions': machine_speed_clean,
            'paper_size_suggestions': paper_size_clean,
            'material_usage_suggestions': material_usage_suggestions,
        })
    
    @action(detail=False, methods=['get'])
    def export_xlsx(self, request):
        """
        Export all paper records to Excel file.
        """
        # Get filtered queryset (respects search, shift, paper_type, and sort parameters)
        queryset = self.get_queryset()
        
        # Filter by date range (Shamsi dates)
        date_from = request.query_params.get('date_from', None)
        date_to = request.query_params.get('date_to', None)
        
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        
        # Create material map for material names
        materials = Material.objects.all()
        material_map = {str(material.id): material.material_name for material in materials}
        
        # Helper function to format material usage
        def format_material_usage(material_usage_json):
            if not material_usage_json:
                return ''
            try:
                material_usage = json.loads(material_usage_json)
                formatted_items = []
                for material_id, data in material_usage.items():
                    if isinstance(data, dict) and 'val' in data:
                        material_name = material_map.get(material_id, f'Material {material_id}')
                        amount = data.get('val', 0)
                        formatted_items.append(f'{material_name}: {amount}')
                return ', '.join(formatted_items)
            except (json.JSONDecodeError, TypeError):
                return material_usage_json
        
        # Helper function to convert datetime to Jalali
        def to_jalali_datetime(dt):
            if not dt:
                return ''
            try:
                jalali_dt = jdatetime.datetime.fromgregorian(datetime=dt)
                return jalali_dt.strftime('%Y-%m-%d %H:%M:%S')
            except:
                return dt.strftime('%Y-%m-%d %H:%M:%S') if dt else ''
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "رکوردهای کاغذ"
        
        # Define headers in Persian
        headers = [
            'شماره رول',
            'خط تولید',
            'تاریخ',
            'زمان شروع نمونه‌گیری',
            'زمان پایان نمونه‌گیری',
            'نام مسئول',
            'شیفت',
            'نوع کاغذ',
            'عرض کاغذ',
            'تعداد پارگی',
            'گراماژ',
            'رطوبت',
            'خاکستر',
            'CUB',
            'پروفایل',
            'شیر غلظت',
            'شیر رقیق‌ساز',
            'دمای سیلندر قبل از سایز پرس',
            'دمای سیلندر بعد از سایز پرس',
            'Burst',
            'MD',
            'CD',
            'CCT 1',
            'CCT 2',
            'CCT 3',
            'CCT 4',
            'CCT 5',
            'RCT 1',
            'RCT 2',
            'RCT 3',
            'RCT 4',
            'RCT 5',
            'زمان پارگی',
            'زمان وقفه در تولید (دقیقه)',
            'علت پارگی/توقف',
            'کلندر اعمال شده',
            'سرعت دستگاه',
            'مصرف مواد',
            'کاربر',
            'تاریخ ایجاد',
            'آخرین بروزرسانی'
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, size=12, color='FFFFFF')
        
        # Write data
        shift_map = {'day': 'روزانه', 'night': 'شبانه'}
        profile_map = {
            '1': '+۱g-',
            '2': '+۲g-',
            '3': '+۳g-',
            '4': '+۴g-',
            '5': 'بیشتر از 5 گرم نوسان سر تا سر کاغذ',
            '+1g': '+۱g-',
            '+2g': '+۲g-',
            '+3g': '+۳g-',
            '+4g': '+۴g-',
            '>5g': 'بیشتر از 5 گرم نوسان سر تا سر کاغذ'
        }
        production_line_map = {2: 'PM2', 3: 'PM3', 4: 'PM4'}
        
        for row_num, paper in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=paper.roll_number or '')
            ws.cell(row=row_num, column=2, value=production_line_map.get(paper.ProductionLine, '') if paper.ProductionLine else '')
            ws.cell(row=row_num, column=3, value=paper.date or '')
            ws.cell(row=row_num, column=4, value=paper.sampling_start_time or '')
            ws.cell(row=row_num, column=5, value=paper.sampling_end_time or '')
            ws.cell(row=row_num, column=6, value=paper.responsible_person_name or '')
            ws.cell(row=row_num, column=7, value=shift_map.get(paper.shift, '') if paper.shift else '')
            ws.cell(row=row_num, column=8, value=paper.PaperType.name if paper.PaperType else '')
            ws.cell(row=row_num, column=9, value=paper.paper_size or '')
            ws.cell(row=row_num, column=10, value=paper.NumberOfTears or '')
            ws.cell(row=row_num, column=11, value=paper.real_grammage or '')
            ws.cell(row=row_num, column=12, value=paper.humidity or '')
            ws.cell(row=row_num, column=13, value=paper.ash_percentage or '')
            ws.cell(row=row_num, column=14, value=paper.cub or '')
            ws.cell(row=row_num, column=15, value=profile_map.get(paper.profile, paper.profile or '') if paper.profile else '')
            ws.cell(row=row_num, column=16, value=paper.density_valve or '')
            ws.cell(row=row_num, column=17, value=paper.diluting_valve or '')
            # Get temperature values from PM_Setting
            pm_settings = paper.pm_settings.all()
            temp_before_values = [str(s.cylinder_temperature_before_press) for s in pm_settings if s.cylinder_temperature_before_press is not None]
            temp_after_values = [str(s.cylinder_temperature_after_press) for s in pm_settings if s.cylinder_temperature_after_press is not None]
            ws.cell(row=row_num, column=18, value=', '.join(temp_before_values) if temp_before_values else '')
            ws.cell(row=row_num, column=19, value=', '.join(temp_after_values) if temp_after_values else '')
            ws.cell(row=row_num, column=20, value=paper.burst_test or '')
            ws.cell(row=row_num, column=21, value=paper.tensile_strength_md or '')
            ws.cell(row=row_num, column=22, value=paper.tensile_strength_cd or '')
            ws.cell(row=row_num, column=23, value=paper.cct1 or '')
            ws.cell(row=row_num, column=24, value=paper.cct2 or '')
            ws.cell(row=row_num, column=25, value=paper.cct3 or '')
            ws.cell(row=row_num, column=26, value=paper.cct4 or '')
            ws.cell(row=row_num, column=27, value=paper.cct5 or '')
            ws.cell(row=row_num, column=28, value=paper.rct1 or '')
            ws.cell(row=row_num, column=29, value=paper.rct2 or '')
            ws.cell(row=row_num, column=30, value=paper.rct3 or '')
            ws.cell(row=row_num, column=31, value=paper.rct4 or '')
            ws.cell(row=row_num, column=32, value=paper.rct5 or '')
            ws.cell(row=row_num, column=33, value=paper.tearing_time or '')
            ws.cell(row=row_num, column=34, value=paper.ProductionDowntime or '')
            ws.cell(row=row_num, column=35, value=paper.CauseOfTearing or '')
            ws.cell(row=row_num, column=36, value='بله' if paper.calender_applied else 'خیر')
            ws.cell(row=row_num, column=37, value=paper.machine_speed or '')
            ws.cell(row=row_num, column=38, value=format_material_usage(paper.material_usage))
            ws.cell(row=row_num, column=39, value=paper.user.username if paper.user else '')
            ws.cell(row=row_num, column=40, value=to_jalali_datetime(paper.created_at))
            ws.cell(row=row_num, column=41, value=to_jalali_datetime(paper.last_updated))
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response with Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Build filename with date range
        date_from = request.query_params.get('date_from', '')
        date_to = request.query_params.get('date_to', '')
        
        if date_from and date_to:
            filename = f'paper-{date_from}-{date_to}.xlsx'
        elif date_from:
            filename = f'paper-{date_from}-.xlsx'
        elif date_to:
            filename = f'paper--{date_to}.xlsx'
        else:
            filename = f'paper-{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        # Set Content-Disposition header with proper encoding
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}'
        
        wb.save(response)
        return response


class ProductionMachineViewSet(viewsets.ModelViewSet):
    """
    ViewSet for ProductionMachine model with CRUD operations.
    """
    queryset = ProductionMachine.objects.all()
    serializer_class = ProductionMachineSerializer
    permission_classes = [AllowAny]