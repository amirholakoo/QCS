"""
Views for pulp app.
"""
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q
from django.http import HttpResponse
from .models import Pulp, pulp_Sampling_Location_names
from .serializers import PulpSerializer, PulpListSerializer
from logs.utils import log_action
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import jdatetime
from django.utils import timezone

from rest_framework.permissions import AllowAny

class PulpViewSet(viewsets.ModelViewSet):
    """
    ViewSet for Pulp model with CRUD operations.
    """
    queryset = Pulp.objects.all()
    serializer_class = PulpSerializer
    permission_classes = [AllowAny]
    
    def get_serializer_class(self):
        """
        Return appropriate serializer based on action.
        """
        if self.action == 'list':
            return PulpListSerializer
        return PulpSerializer
    
    def get_queryset(self):
        """
        Filter queryset based on query parameters.
        """
        queryset = Pulp.objects.all()
        
        # Search functionality
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(roll_number__icontains=search)
            )
        
        # Sorting
        sort_by = self.request.query_params.get('sort_by', '-created_at')
        if sort_by:
            queryset = queryset.order_by(sort_by)
        
        return queryset
    
    def perform_create(self, serializer):
        """
        Create pulp record and log the action.
        """
        pulp = serializer.save()
        
        # Log action if user is authenticated
        if self.request.user.is_authenticated:
            try:
                log_action(self.request.user.username, 'Pulp', 'create',[{'roll_number':pulp.roll_number}])
            except:
                pass
    
    def perform_update(self, serializer):
        """
        Update pulp record and log the action.
        """
        serializer.save()
        
        # Log action if user is authenticated
        # if self.request.user.is_authenticated:
        #     try:
        #         log_action(self.request.user.username, 'Pulp', 'edit')
        #     except:
        #         pass
    
    def perform_destroy(self, instance):
        """
        Delete pulp record and log the action.
        """
        instance.delete()
        
        # Log action if user is authenticated
        if self.request.user.is_authenticated:
            try:
                log_action(self.request.user.username, 'Pulp', 'delete',[{'roll_number':instance.roll_number}])
            except:
                pass
    
    @action(detail=False, methods=['get'])
    def export_xlsx(self, request):
        """
        Export all pulp records to Excel file.
        """
        # Get filtered queryset (respects search and sort parameters)
        queryset = self.get_queryset()
        
        # Filter by date range (convert Shamsi to Gregorian for created_at filtering)
        date_from = request.query_params.get('date_from', None)
        date_to = request.query_params.get('date_to', None)
        
        if date_from:
            try:
                # Convert Shamsi date to Gregorian datetime
                year, month, day = map(int, date_from.split('-'))
                jalali_date = jdatetime.date(year, month, day)
                gregorian_date = jalali_date.togregorian()
                # Start of day
                start_datetime = timezone.make_aware(
                    datetime.combine(gregorian_date, datetime.min.time())
                )
                queryset = queryset.filter(created_at__gte=start_datetime)
            except (ValueError, AttributeError):
                pass  # Invalid date format, skip filter
        
        if date_to:
            try:
                # Convert Shamsi date to Gregorian datetime
                year, month, day = map(int, date_to.split('-'))
                jalali_date = jdatetime.date(year, month, day)
                gregorian_date = jalali_date.togregorian()
                # End of day
                end_datetime = timezone.make_aware(
                    datetime.combine(gregorian_date, datetime.max.time())
                )
                queryset = queryset.filter(created_at__lte=end_datetime)
            except (ValueError, AttributeError):
                pass  # Invalid date format, skip filter
        
        # Helper function to convert datetime to Jalali
        def to_jalali_datetime(dt):
            if not dt:
                return ''
            try:
                jalali_dt = jdatetime.datetime.fromgregorian(datetime=dt)
                return jalali_dt.strftime('%Y-%m-%d %H:%M:%S')
            except:
                return dt.strftime('%Y-%m-%d %H:%M:%S') if dt else ''
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "نمونه‌های خمیر"
        
        # Define headers in Persian
        headers = [
            'شماره رول',
            'خط تولید',
            'زمان نمونه‌گیری پایین',
            'کانس خمیر پایین',
            'فیلتر آب پایین',
            'فرینس خمیر پایین',
            'pH پایین',
            'دمای خمیر پایین',
            'غلظت هدباکس بالا',
            'فیلتر آب بالا',
            'فرینس هدباکس بالا',
            'pH بالا',
            'دمای خمیر بالا',
            'کانس حوض 8',
            'کانس کردان',
            'کانس تیکنر',
            'تاریخ ایجاد',
            'آخرین بروزرسانی'
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, size=12, color='FFFFFF')
        
        # Write data
        production_line_map = {0: 'مشترک', 2: 'PM2', 3: 'PM3', 4: 'PM4'}
        
        for row_num, pulp in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=pulp.roll_number or '')
            ws.cell(row=row_num, column=2, value=production_line_map.get(pulp.ProductionLine, '') if pulp.ProductionLine is not None else '')
            ws.cell(row=row_num, column=3, value=pulp.lower_sampling_time or '')
            ws.cell(row=row_num, column=4, value=pulp.downpulpcount or '')
            ws.cell(row=row_num, column=5, value=pulp.lower_water_filter or '')
            ws.cell(row=row_num, column=6, value=pulp.lower_headbox_freeness or '')
            ws.cell(row=row_num, column=7, value=pulp.lower_ph or '')
            ws.cell(row=row_num, column=8, value=pulp.lower_pulp_temperature or '')
            ws.cell(row=row_num, column=9, value=pulp.upper_headbox_consistency or '')
            ws.cell(row=row_num, column=10, value=pulp.upper_water_filter or '')
            ws.cell(row=row_num, column=11, value=pulp.upper_headbox_freeness or '')
            ws.cell(row=row_num, column=12, value=pulp.upper_ph or '')
            ws.cell(row=row_num, column=13, value=pulp.upper_pulp_temperature or '')
            ws.cell(row=row_num, column=14, value=pulp.pond8_consistency or '')
            ws.cell(row=row_num, column=15, value=pulp.curtain_consistency or '')
            ws.cell(row=row_num, column=16, value=pulp.thickener_consistency or '')
            ws.cell(row=row_num, column=17, value=to_jalali_datetime(pulp.created_at))
            ws.cell(row=row_num, column=18, value=to_jalali_datetime(pulp.last_updated))
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create HTTP response with Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Build filename with date range
        date_from = request.query_params.get('date_from', '')
        date_to = request.query_params.get('date_to', '')
        
        if date_from and date_to:
            filename = f'pulp-{date_from}-{date_to}.xlsx'
        elif date_from:
            filename = f'pulp-{date_from}-.xlsx'
        elif date_to:
            filename = f'pulp--{date_to}.xlsx'
        else:
            filename = f'pulp-{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        # Set Content-Disposition header with proper encoding
        response['Content-Disposition'] = f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}'
        
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'])
    def location_names(self, request):
        """
        Get all unique location names for suggestions.
        """
        location_names = pulp_Sampling_Location_names.objects.all().order_by('title')
        names = [{'id': loc.id, 'title': loc.title} for loc in location_names]
        return Response(names)